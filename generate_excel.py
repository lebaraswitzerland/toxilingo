import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define categories and data generator
scenarios = [
    # --- ESPAGNOL (MEDELLIN / PAISA / NSFW / SLANG) [20 Videos] ---
    {
        "ID": "VIDEO_01",
        "Langue": "Espagnol",
        "Theme": "NSFW / Double Entendre",
        "Question": "Como se dice 'can I eat cake' en español?",
        "Reponse": "Como culo? You eat ass? No wonder your breath smells like shit!",
        "Prompt_Image": "Raw smartphone selfie, low angle shot from chest level, 22yo brunette girl with slight facial redness, natural skin texture, visible pores, wearing a oversized grey hoodie in a messy bedroom, natural daylight, photorealistic, 35mm lens effect --no smooth skin, no anime, no 3d render",
        "Prompt_Veo3": "The girl takes a bite of a pastry, freezes in absolute shock, eyes wide, mouth full, chewing awkwardly as tears of embarrassment well up in her eyes, raw camera shake",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_02",
        "Langue": "Espagnol",
        "Theme": "NSFW / Misunderstanding",
        "Question": "Como se dice 'I eat pie' en español?",
        "Reponse": "Yo como pies! You said I eat feet! Please, can you eat mine perrita?",
        "Prompt_Image": "Selfie from low angle, 23yo blonde girl with wet hair, slight mascara smudge under eyes, plain white t-shirt, sitting in a car, sunny afternoon lighting, authentic iPhone front camera quality, detailed skin flaws",
        "Prompt_Veo3": "Girl looks at her phone in disbelief, starts crying hysterically while gasping for air, shaking her head no in extreme embarrassment",
        "Emotion": "Crying"
    },
    {
        "ID": "VIDEO_03",
        "Langue": "Espagnol",
        "Theme": "NSFW / Car Smell",
        "Question": "Como se dice 'my car smells like eggs' en español?",
        "Reponse": "Mi coochie huele a huevos! Go wash your pussy pendeja!",
        "Prompt_Image": "Low angle smartphone selfie, 24yo girl with blonde ponytail, wearing tank top, ceiling fan spinning in background of bedroom, authentic unpolished look, high resolution skin texture with freckles",
        "Prompt_Veo3": "Girl bursts into uncontrollable laughter that turns into ugly crying, throwing her head back in complete defeat",
        "Emotion": "Laughing"
    },
    {
        "ID": "VIDEO_04",
        "Langue": "Espagnol (Paisa / Medellin)",
        "Theme": "Medellin Slang Fail",
        "Question": "How do I ask for a hot chocolate in Medellin?",
        "Reponse": "Dame el choco-ano tía! Congrats gringo, you just asked the waitress to eat her asshole with cheese!",
        "Prompt_Image": "Low angle view of a 21yo girl sitting on a balcony in Medellin, casual summer clothes, afternoon glow, realistic skin imperfections, slight sweat on forehead, unedited smartphone photo style",
        "Prompt_Veo3": "Girl drops her jaw, covers her mouth in horror, eyes darting left and right in sheer panic",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_05",
        "Langue": "Espagnol",
        "Theme": "Slang / Insult",
        "Question": "Como se dice 'I am cool' en Colombia?",
        "Reponse": "Soy un gonorrea de primera! Perfect, now everyone knows you are a walking venereal disease!",
        "Prompt_Image": "Selfie angle, 22yo girl in a cozy kitchen, wearing pajamas, messy bun, natural lighting, realistic facial skin with subtle blemishes and texture, raw camera shot",
        "Prompt_Veo3": "Girl squints at screen, winces in disgust, face contorting into an awkward grimace as she bites her lip",
        "Emotion": "Disgusted"
    },
    {
        "ID": "VIDEO_06",
        "Langue": "Espagnol",
        "Theme": "NSFW / Food",
        "Question": "How do I say 'I love cucumber' in Spanish?",
        "Reponse": "Me encanta la verga en ensalada! Outstanding, you just declared your love for raw dick in your salad!",
        "Prompt_Image": "Low angle phone camera perspective, 23yo girl lying on a bed with a white pillow, wearing a casual sweater, authentic acne scars visible, soft realistic skin",
        "Prompt_Veo3": "Girl snorts out loud trying to hold back a laugh, blushes red, and buries her face in her hands",
        "Emotion": "Laughing"
    },
    {
        "ID": "VIDEO_07",
        "Langue": "Espagnol",
        "Theme": "Insult / Relationship",
        "Question": "How do I compliment my Colombian girlfriend's eyes?",
        "Reponse": "Tienes ojos de perra con rabia! Aww, telling her she has eyes of a rabid bitch is so romantic!",
        "Prompt_Image": "Selfie from below, 24yo woman with dark curly hair, green eyes, wearing denim jacket, outdoor urban cafe background, unedited smartphone camera grain",
        "Prompt_Veo3": "Girl shakes her head slowly with tears streaming down her cheeks, mouthing 'why' to the camera",
        "Emotion": "Crying"
    },
    {
        "ID": "VIDEO_08",
        "Langue": "Espagnol",
        "Theme": "Slang / Gringo Fail",
        "Question": "Como se dice 'I am embarrassed' in Spanish?",
        "Reponse": "Estoy embarazada! Wow congrats bro, a gringo guy getting pregnant is a medical miracle!",
        "Prompt_Image": "Low angle selfie, 22yo girl in a living room, wearing headphones around neck, natural light, skin texture with minor redness, raw authentic look",
        "Prompt_Veo3": "Girl facepalms hard, leaving a red mark on her forehead, while wheezing from silent laughter",
        "Emotion": "Laughing"
    },
    {
        "ID": "VIDEO_09",
        "Langue": "Espagnol",
        "Theme": "NSFW / Taxi",
        "Question": "How do I tell the taxi driver 'take me to the beach'?",
        "Reponse": "Llévame al bicho cabrón! Great job, you just asked a scary local taxi driver to take you by his dick!",
        "Prompt_Image": "Low angle view inside a car, 25yo woman passenger seat selfie, seatbelt across chest, natural afternoon light, realistic skin, no makeup look",
        "Prompt_Veo3": "Girl's eyes widen in terror, she looks out the window frantically, swallowing hard in panic",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_10",
        "Langue": "Espagnol",
        "Theme": "Insult / Intelligence",
        "Question": "Como se dice 'I don't speak Spanish well'?",
        "Reponse": "Soy retrasado mental no me peguen! Exactly right, begging locals not to beat you up for being stupid!",
        "Prompt_Image": "Low angle photo, 23yo girl sitting on floor against a wall, oversized t-shirt, messy hair, ultra-detailed skin textures, natural home lighting",
        "Prompt_Veo3": "Girl lets out a dramatic sob, wiping her nose with her sleeve, looking utterly defeated",
        "Emotion": "Crying"
    },
    {
        "ID": "VIDEO_11",
        "Langue": "Espagnol",
        "Theme": "NSFW / Fruit",
        "Question": "How do I ask for a papaya at the market?",
        "Reponse": "Déjame manosear tu papaya mamacita! Perfect, sexual harassment charges in 3... 2... 1...",
        "Prompt_Image": "Selfie from below, 22yo girl at a wooden desk with a laptop, casual wear, soft natural lighting, realistic facial pores and subtle lip dryness",
        "Prompt_Veo3": "Girl cringes severely, pulling her chin back into a double chin posture while grimacing in horror",
        "Emotion": "Disgusted"
    },
    {
        "ID": "VIDEO_12",
        "Langue": "Espagnol",
        "Theme": "Insult / Appearance",
        "Question": "Como se dice 'I look handsome today'?",
        "Reponse": "Parezco un chimpancé bañado! Spot on, you look like a freshly washed gorilla!",
        "Prompt_Image": "Low angle phone shot, 24yo girl holding a coffee mug, cozy hoodie, morning sunlight, ultra realistic skin details, no filter",
        "Prompt_Veo3": "Girl wheezes out a laugh, spilling a bit of coffee, frantically wiping it while giggling",
        "Emotion": "Laughing"
    },
    {
        "ID": "VIDEO_13",
        "Langue": "Espagnol",
        "Theme": "NSFW / Body",
        "Question": "How to say 'my back hurts' in Spanish?",
        "Reponse": "Me duele el culo de tanto azotar! Wow, bragging about getting your ass clapped too hard!",
        "Prompt_Image": "Low angle selfie, 21yo girl lying in bed under duvet, morning light, natural skin texture, authentic lazy morning vibe",
        "Prompt_Veo3": "Girl buries her face into the pillow screaming in laughter, feet kicking in the air",
        "Emotion": "Laughing"
    },
    {
        "ID": "VIDEO_14",
        "Langue": "Espagnol",
        "Theme": "Paisa Slang / Party",
        "Question": "How to ask 'where is the party' in Medellin?",
        "Reponse": "Dónde hay prepagos baratas socio? Nice! You just asked a random police officer where to buy cheap prostitutes!",
        "Prompt_Image": "Selfie from chest height looking up, 23yo girl with hoop earrings, dim night room with warm lamp light, realistic skin grain, unpolished portrait",
        "Prompt_Veo3": "Girl's mouth hangs open in total disbelief, shaking her head vigorously 'no no no'",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_15",
        "Langue": "Espagnol",
        "Theme": "NSFW / Pet",
        "Question": "Como se dice 'I love my cat'?",
        "Reponse": "Me chupo mi propia chocha! Incredible, claiming you can lick your own vagina!",
        "Prompt_Image": "Low angle smartphone selfie, 22yo girl petting a cat in background, wearing tank top, natural lighting, realistic facial imperfections",
        "Prompt_Veo3": "Girl turns bright red, covers her eyes with both hands, crying tears of laughter",
        "Emotion": "Crying"
    },
    {
        "ID": "VIDEO_16",
        "Langue": "Espagnol",
        "Theme": "Insult / Restaurant",
        "Question": "How do I tell the waiter the food was good?",
        "Reponse": "Esta mierda es comestible milagro! Complimenting the chef that his shit is surprisingly edible!",
        "Prompt_Image": "Low angle photo, 24yo girl sitting at dining table, simple clothes, realistic skin tone, soft evening lamp glow, zero retouching",
        "Prompt_Veo3": "Girl winces in awkward cringe, giving a slow, hesitant thumbs up to the camera",
        "Emotion": "Disgusted"
    },
    {
        "ID": "VIDEO_17",
        "Langue": "Espagnol",
        "Theme": "NSFW / Weather",
        "Question": "How do I say 'it's very hot today'?",
        "Reponse": "Estoy más caliente que una perra en celo! Beautiful, telling locals you are horny like a bitch in heat!",
        "Prompt_Image": "Low angle camera selfie, 23yo brunette with hair tied in a messy bun, fan blowing air, summer clothing, realistic skin texture with sweat sheen",
        "Prompt_Veo3": "Girl wipes sweat from her forehead, fanatically fanning herself while groaning in embarrassment",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_18",
        "Langue": "Espagnol",
        "Theme": "Insult / Shopping",
        "Question": "How do I ask 'is there a discount'?",
        "Reponse": "Soy tacaño y podrido téngame lástima! Asking for mercy because you are a rotten cheapskate!",
        "Prompt_Image": "Low angle selfie, 22yo girl wearing cardigan, holding shopping bag, indoor light, natural skin detail, realistic 35mm photograph",
        "Prompt_Veo3": "Girl rolls her eyes dramatically, sighing out loud with a pained expression",
        "Emotion": "Angry"
    },
    {
        "ID": "VIDEO_19",
        "Langue": "Espagnol",
        "Theme": "NSFW / Drinks",
        "Question": "How do I order a cold milk?",
        "Reponse": "Dame leche de tu manguera papi! Asking the bartender to feed you milk straight from his hose!",
        "Prompt_Image": "Low angle smartphone shot, 25yo woman in kitchen holding a glass, natural daylight, photorealistic skin with small freckles",
        "Prompt_Veo3": "Girl chokes on her drink, coughing frantically with wide eyes",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_20",
        "Langue": "Espagnol",
        "Theme": "Insult / Gym",
        "Question": "Como se dice 'I workout every day'?",
        "Reponse": "Inyecto esteroides y sigo fofo! Telling everyone at the gym you inject juice and stay soft!",
        "Prompt_Image": "Low angle post-workout selfie, 23yo girl in sports bra, towel on shoulders, gym room background, sweaty natural skin texture",
        "Prompt_Veo3": "Girl flexes sarcastically, then breaks down into uncontrollable giggles",
        "Emotion": "Laughing"
    },

    # --- JAPONAIS (YAKUZA / SENSEI / OTAKU) [10 Videos] ---
    {
        "ID": "VIDEO_21",
        "Langue": "Japonais",
        "Theme": "Yakuza / Extreme Honor",
        "Question": "How do I say 'excuse me' in Tokyo?",
        "Reponse": "Ketsu wo nameasete kudasai! Congrats otaku, you just begged a Yakuza boss to let you lick his ass!",
        "Prompt_Image": "Low angle selfie, 22yo girl with black straight hair, wearing Japanese streetwear hoodie, Tokyo room background, realistic skin texture, authentic lighting",
        "Prompt_Veo3": "Girl bows repeatedly in panic, tears welling up as she trembles in fear",
        "Emotion": "Crying"
    },
    {
        "ID": "VIDEO_22",
        "Langue": "Japonais",
        "Theme": "Anime Cliché Fail",
        "Question": "How do I call a cute girl 'senpai'?",
        "Reponse": "Ore no baka chinchin wo miro! Perfect, screaming 'look at my stupid penis' at a stranger!",
        "Prompt_Image": "Low angle phone camera, 21yo girl sitting near anime posters in bedroom, natural light, skin texture with minor flaws, authentic front camera shot",
        "Prompt_Veo3": "Girl covers her face with an anime manga book, whimpering in shame",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_23",
        "Langue": "Japonais",
        "Theme": "Ramen Shop Fail",
        "Question": "How do I say 'this soup is delicious'?",
        "Reponse": "Kono shiru wa opai no aji ga suru! Telling the old chef his ramen tastes like breast milk!",
        "Prompt_Image": "Low angle selfie, 23yo girl holding chopsticks, warm wood lighting, realistic skin details, slight lip gloss, unedited photo",
        "Prompt_Veo3": "Girl drops chopsticks on table, staring blankly with open mouth in utter paralysis",
        "Emotion": "Disgusted"
    },
    {
        "ID": "VIDEO_24",
        "Langue": "Japonais",
        "Theme": "Sensei Aggression",
        "Question": "How do I ask 'can you speak English'?",
        "Reponse": "Eigo hanase baka gaijin domo! Shouting 'speak English you stupid foreigner pigs!'",
        "Prompt_Image": "Low angle perspective, 24yo girl at study desk with textbooks, realistic skin tone, soft lamp light, raw smartphone quality",
        "Prompt_Veo3": "Girl slams desk with her palm, looking furious before breaking into tears",
        "Emotion": "Angry"
    },
    {
        "ID": "VIDEO_25",
        "Langue": "Japonais",
        "Theme": "Otaku / Maid Cafe",
        "Question": "How to order tea at a Maid Cafe?",
        "Reponse": "O-kyaku sama noパンツ wo kudasai! Asking the cute maid to hand over her underwear!",
        "Prompt_Image": "Low angle selfie, 22yo girl with bangs, oversized cardigan, neon street lights from window, realistic skin texture with freckles",
        "Prompt_Veo3": "Girl blushes deep red, fanning her face frantically with both hands",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_26",
        "Langue": "Japonais",
        "Theme": "Taxi / GPS",
        "Question": "How to say 'turn right here'?",
        "Reponse": "Koko de seppuku shimasu! Informing the cab driver you will commit ritual suicide right here!",
        "Prompt_Image": "Low angle selfie in backseat of car, 23yo girl, nighttime city lights reflection on face, photorealistic skin grain",
        "Prompt_Veo3": "Girl gulps loudly, eyes wide open staring at the driver in terror",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_27",
        "Langue": "Japonais",
        "Theme": "Polite / Honorific",
        "Question": "How do I say 'thank you for the meal'?",
        "Reponse": "Gochisousama! Now pay double because you sounded like a samurai from 1500!",
        "Prompt_Image": "Low angle selfie, 24yo girl holding tea cup, natural light, realistic skin, cozy sweater",
        "Prompt_Veo3": "Girl smirks awkwardly, tilting her head with a pained smile",
        "Emotion": "Disgusted"
    },
    {
        "ID": "VIDEO_28",
        "Langue": "Japonais",
        "Theme": "NSFW / Onsen",
        "Question": "How do I ask where the hot spring is?",
        "Reponse": "Hadaka de chinchin dance ga taiketsu! Inviting elderly locals to a naked penis dance duel!",
        "Prompt_Image": "Low angle selfie, 22yo girl with towel wrapped hair, bathroom background, humid lighting, realistic wet skin texture",
        "Prompt_Veo3": "Girl covers face with towel, groaning hysterically in embarrassment",
        "Emotion": "Crying"
    },
    {
        "ID": "VIDEO_29",
        "Langue": "Japonais",
        "Theme": "Subway / Traffic",
        "Question": "How to say 'the train is full'?",
        "Reponse": "Chikan ga tatakatte iru! Screaming 'the molesters are fighting each other!' on train",
        "Prompt_Image": "Low angle phone selfie, 23yo girl with scarf, train window background, authentic smartphone camera noise, realistic skin",
        "Prompt_Veo3": "Girl gasps violently, ducking her head down in horror",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_30",
        "Langue": "Japonais",
        "Theme": "Cat / Cute",
        "Question": "How to say 'kawaii cat'?",
        "Reponse": "Neko no kintama sugoi! Praising the magnificence of a cat's testicles!",
        "Prompt_Image": "Low angle selfie, 21yo girl holding a plushie, bedroom background, soft realistic lighting, unpolished raw look",
        "Prompt_Veo3": "Girl tries to hold a laugh, nose snorting, bursting into giggles",
        "Emotion": "Laughing"
    },

    # --- ALLEMAND (AGRESSIF / MILITAIRE / ABSURDE) [10 Videos] ---
    {
        "ID": "VIDEO_31",
        "Langue": "Allemand",
        "Theme": "Aggressive Cliché",
        "Question": "How do I say 'good morning' politely?",
        "Reponse": "HALT MAUL UND ARBEITE SCHWEIN! Shouting 'SHUT UP AND WORK YOU SWINE!' at 7 AM!",
        "Prompt_Image": "Low angle selfie, 23yo blonde girl in black turtleneck, overcast daylight, realistic pale skin texture, raw photo",
        "Prompt_Veo3": "Girl flinches backward as if slapped by sound, holding her heart in shock",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_32",
        "Langue": "Allemand",
        "Theme": "NSFW / Beer",
        "Question": "How do I order one beer in Munich?",
        "Reponse": "Ein Bier und spritz mir ins Gesicht! Ordering one beer with a side of facial cumshot!",
        "Prompt_Image": "Low angle selfie, 24yo girl holding a pretzel, pub background, dim warm light, realistic skin grain",
        "Prompt_Veo3": "Girl chokes on her breath, spit-taking into her hands in complete horror",
        "Emotion": "Crying"
    },
    {
        "ID": "VIDEO_33",
        "Langue": "Allemand",
        "Theme": "Bakery / Bread",
        "Question": "How do I ask for fresh bread?",
        "Reponse": "Ich möchte deine Oma essen! Asking the baker if you can eat his grandmother!",
        "Prompt_Image": "Low angle selfie, 22yo girl near kitchen counter, realistic skin imperfections, authentic daylight",
        "Prompt_Veo3": "Girl shakes her head 'no', eyebrows furrowed in utter disgust",
        "Emotion": "Disgusted"
    },
    {
        "ID": "VIDEO_34",
        "Langue": "Allemand",
        "Theme": "Insult / Bureaucracy",
        "Question": "How to say 'I have a passport'?",
        "Reponse": "Ich bin ein illegaler Kartoffelkopf! Declaring yourself an illegal potato head!",
        "Prompt_Image": "Low angle photo, 25yo woman at office desk, realistic skin texture, plain shirt, raw camera look",
        "Prompt_Veo3": "Girl bursts into wheezing laugh, rubbing her eyes under glasses",
        "Emotion": "Laughing"
    },
    {
        "ID": "VIDEO_35",
        "Langue": "Allemand",
        "Theme": "NSFW / Sausage",
        "Question": "How do I say 'I love Bratwurst'?",
        "Reponse": "Ich stecke Wurst in meine Ritze! Telling the butcher you shove sausages up your crack!",
        "Prompt_Image": "Low angle selfie, 22yo girl in coat, outdoor street background, cold lighting, realistic skin detail",
        "Prompt_Veo3": "Girl covers her mouth, shaking in silent tearful laughter",
        "Emotion": "Crying"
    },
    {
        "ID": "VIDEO_36",
        "Langue": "Allemand",
        "Theme": "Taxi / Speed",
        "Question": "How do I tell Autobahn driver to go faster?",
        "Reponse": "Fahr schneller du lahme Ente oder ich kotze! 'Drive faster you lame duck or I puke!'",
        "Prompt_Image": "Low angle selfie inside car, 23yo girl, motion blur background, realistic skin, front camera style",
        "Prompt_Veo3": "Girl clings to grab handle, eyes wild with panic",
        "Emotion": "Angry"
    },
    {
        "ID": "VIDEO_37",
        "Langue": "Allemand",
        "Theme": "Romance / Dating",
        "Question": "How do I say 'you are beautiful'?",
        "Reponse": "Du riechst wie eine alte Wurst! Complimenting her that she smells like an old sausage!",
        "Prompt_Image": "Low angle selfie, 24yo girl with red lipstick, date night outfit, cozy restaurant background, realistic skin",
        "Prompt_Veo3": "Girl looks down in shame, biting her lip awkwardly",
        "Emotion": "Disgusted"
    },
    {
        "ID": "VIDEO_38",
        "Langue": "Allemand",
        "Theme": "Hotel / Room",
        "Question": "How to say 'the bed is comfortable'?",
        "Reponse": "Das Bett ist voll mit Flöhen! Telling the hotel manager the bed is full of fleas!",
        "Prompt_Image": "Low angle selfie on hotel bed, 22yo girl, white sheets, realistic skin texture, soft lighting",
        "Prompt_Veo3": "Girl jumps off bed in horror, scratching arms frantically",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_39",
        "Langue": "Allemand",
        "Theme": "Doctor / Health",
        "Question": "How do I say 'I have a headache'?",
        "Reponse": "Mein Gehirn schmilzt hilfe! Screaming 'my brain is melting help me!' at the pharmacist",
        "Prompt_Image": "Low angle selfie, 23yo girl holding ice pack to head, realistic pale skin texture, raw photo",
        "Prompt_Veo3": "Girl groans dramatically, letting head fall back onto sofa",
        "Emotion": "Crying"
    },
    {
        "ID": "VIDEO_40",
        "Langue": "Allemand",
        "Theme": "Cat / Pet",
        "Question": "How do I say 'cute dog'?",
        "Reponse": "Böser Hund beiß mich nicht! Yelling 'evil dog don't bite me!' while trembling",
        "Prompt_Image": "Low angle selfie, 21yo girl in park, sunny day, realistic skin freckles, unedited photo",
        "Prompt_Veo3": "Girl giggles uncontrollably, covering mouth with sweater sleeve",
        "Emotion": "Laughing"
    },

    # --- FRANÇAIS (SNOB ARROGANT / MEPRIS D'ACCENT FOR US) [10 Videos] ---
    {
        "ID": "VIDEO_41",
        "Langue": "Français",
        "Theme": "Arrogant Snob",
        "Question": "How do I say 'can I get a croissant' in Paris?",
        "Reponse": "Tais-toi américain ton accent me donne la nausée! Shouting 'Shut up American, your accent makes me vomit!'",
        "Prompt_Image": "Low angle selfie, 23yo French brunette girl with beret, Paris cafe terrace background, realistic skin tone with subtle nose shine, 35mm photograph",
        "Prompt_Veo3": "Girl rolls eyes into back of head, letting out a dramatic French sigh of disgust",
        "Emotion": "Disgusted"
    },
    {
        "ID": "VIDEO_42",
        "Langue": "Français",
        "Theme": "NSFW / Bakery",
        "Question": "How do I ask for a warm baguette?",
        "Reponse": "Met ta baguette dans mes fesses! Asking the baker to shove his baguette up your butt!",
        "Prompt_Image": "Low angle selfie, 22yo girl holding paper bag, kitchen background, morning light, realistic skin textures",
        "Prompt_Veo3": "Girl gasps loudly, dropping her jaw in absolute disbelief",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_43",
        "Langue": "Français",
        "Theme": "Wine / Restaurant",
        "Question": "How to say 'this red wine is good'?",
        "Reponse": "Ce jus de pisse est passable! Telling the sommelier his wine tastes like passable urine!",
        "Prompt_Image": "Low angle selfie, 24yo girl with wine glass, dim restaurant, realistic skin texture, cozy night shot",
        "Prompt_Veo3": "Girl spits wine back into glass, coughing while trying to laugh",
        "Emotion": "Laughing"
    },
    {
        "ID": "VIDEO_44",
        "Langue": "Français",
        "Theme": "NSFW / Kiss",
        "Question": "How do I ask a girl for a kiss?",
        "Reponse": "Lèche ma langue comme un crapaud! Asking her to lick your tongue like a toad!",
        "Prompt_Image": "Low angle selfie, 23yo girl with red sweater, park background, autumn light, realistic skin",
        "Prompt_Veo3": "Girl sticks tongue out in disgust, cringing severely",
        "Emotion": "Disgusted"
    },
    {
        "ID": "VIDEO_45",
        "Langue": "Français",
        "Theme": "Taxi / Paris",
        "Question": "How to ask 'are we almost there'?",
        "Reponse": "Avance connard je suis en retard! Calling the Parisian taxi driver an asshole for being slow!",
        "Prompt_Image": "Low angle selfie in backseat of car, 25yo woman, rain on window, realistic skin texture, unedited look",
        "Prompt_Veo3": "Girl covers ears, terrified of driver's response",
        "Emotion": "Shocked"
    },
    {
        "ID": "VIDEO_46",
        "Langue": "Français",
        "Theme": "Fashion / Shop",
        "Question": "How to say 'this dress is pretty'?",
        "Reponse": "Ce torchon me fait ressembler à un sac! Calling the designer dress a dishcloth that makes you look like a sack!",
        "Prompt_Image": "Low angle selfie in mirror/shop, 22yo girl, stylish coat, indoor shop lighting, realistic skin texture",
        "Prompt_Veo3": "Girl shakes head dramatically, facepalming into hands",
        "Emotion": "Crying"
    },
    {
        "ID": "VIDEO_47",
        "Langue": "Français",
        "Theme": "NSFW / Cheese",
        "Question": "How to order smelly cheese?",
        "Reponse": "Ça sent les pieds de mon grand-père! Informing the shopkeeper it smells like your grandpa's feet!",
        "Prompt_Image": "Low angle selfie, 24yo girl holding cheese block, kitchen background, natural light, realistic skin",
        "Prompt_Veo3": "Girl pinches nose, eyes watering in laughter",
        "Emotion": "Laughing"
    },
    {
        "ID": "VIDEO_48",
        "Langue": "Français",
        "Theme": "Eiffel Tower",
        "Question": "How to ask where the Eiffel Tower is?",
        "Reponse": "Où est la grande tour en ferraille moche? Asking for the ugly iron tower!",
        "Prompt_Image": "Low angle selfie outdoors, 21yo girl, windy day, hair across face, realistic skin freckles",
        "Prompt_Veo3": "Girl laughs so hard she cries, holding her stomach",
        "Emotion": "Laughing"
    },
    {
        "ID": "VIDEO_49",
        "Langue": "Français",
        "Theme": "Hotel / Service",
        "Question": "How to ask for more towels?",
        "Reponse": "Donne moi du chiffon sale! Asking room service for dirty rags!",
        "Prompt_Image": "Low angle selfie, 23yo girl in bathrobe, hotel room background, soft realistic lighting",
        "Prompt_Veo3": "Girl looks confused, tilting head in frustration",
        "Emotion": "Angry"
    },
    {
        "ID": "VIDEO_50",
        "Langue": "Français",
        "Theme": "NSFW / Love",
        "Question": "How do I say 'I love you my sweetie'?",
        "Reponse": "Je t'aime mon petit boudin gras! Calling your lover a fat little blood sausage!",
        "Prompt_Image": "Low angle selfie, 22yo girl in bed, warm cozy light, realistic skin, front camera unedited shot",
        "Prompt_Veo3": "Girl bursts into tears of laughter, hiding under blanket",
        "Emotion": "Laughing"
    }
]

# Convert to DataFrame
df = pd.DataFrame(scenarios)

# Excel File Export with Styling
file_name = "toxilingo_50_scripts.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "50 Scripts Viraux ToxiLingo"

# Header styles
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

headers = ["ID", "Langue Cible", "Thème / Niche", "Question Humain", "Réponse App Toxique", "Prompt Image (Flux/MJ)", "Prompt Animation (Veo 3)", "Émotion"]
ws.append(headers)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align

# Insert rows
for row_num, item in enumerate(scenarios, 2):
    row_data = [
        item["ID"],
        item["Langue"],
        item["Theme"],
        item["Question"],
        item["Reponse"],
        item["Prompt_Image"],
        item["Prompt_Veo3"],
        item["Emotion"]
    ]
    ws.append(row_data)
    
    # Apply cell styling
    for col_num in range(1, len(row_data) + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.border = thin_border
        if col_num in [1, 2, 8]:
            cell.alignment = center_align
        else:
            cell.alignment = left_align

# Column widths
col_widths = {
    "A": 12,
    "B": 15,
    "C": 22,
    "D": 35,
    "E": 45,
    "F": 55,
    "G": 55,
    "H": 15
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze header row
ws.freeze_panes = "A2"

# Save workbook
wb.save(file_name)
print(f"✅ Excel file '{file_name}' generated successfully with 50 viral scenarios!")
